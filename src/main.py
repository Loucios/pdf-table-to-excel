import re
from pathlib import Path

import pandas as pd
import pdfplumber
import tabula
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


INPUT_DIR = Path("input")
OUTPUT_DIR = Path("output")

# ------------------------------------------------------------
# НАСТРОЙКА ФОРМАТА НАЗВАНИЙ ТАБЛИЦ
# ------------------------------------------------------------
# Что поддерживает:
# Таблица 1 - Текст
# Таблица 1.2 - Текст
# Таблица 1.2.3 - Текст
# Таблица А.1 - Текст
# Таблица Б.2.1 - Текст
#
# Важно:
# [А-ЯA-Z] позволяет искать как русские буквы, так и латинские.
# [-–—] позволяет искать дефис, короткое тире и длинное тире.
# ------------------------------------------------------------

TABLE_NUMBER_PATTERN = r"(?:\d+(?:\.\d+)*|[А-ЯA-Z]\.\d+(?:\.\d+)*)"

TABLE_TITLE_PATTERN = re.compile(
    rf"(Таблица\s+{TABLE_NUMBER_PATTERN}\s*[-–—]\s*.+)",
    re.IGNORECASE
)


def safe_sheet_name(name: str) -> str:
    """
    Excel ограничивает имя листа 31 символом
    и запрещает некоторые символы.
    """
    invalid_chars = r'[]:*?/\\'
    for char in invalid_chars:
        name = name.replace(char, " ")

    name = re.sub(r"\s+", " ", name).strip()
    return name[:31] if name else "Table"


def normalize_header_value(value):
    """
    Нормализует значение для сравнения заголовков:
    убирает пробелы, переносы строк, приводит к нижнему регистру.
    """
    if pd.isna(value):
        return ""

    text = str(value)
    text = text.replace("\n", " ")
    text = text.replace("\xa0", " ")
    text = text.replace("\u202f", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()

    return text


def is_repeated_header_row(row, columns) -> bool:
    """
    Проверяет, является ли первая строка таблицы повтором заголовка колонок.

    Это нужно, когда tabula на каждой странице вытаскивает продолжение таблицы
    вместе с повторяющейся шапкой.
    """
    row_values = [normalize_header_value(value) for value in row]
    col_values = [normalize_header_value(value) for value in columns]

    if len(row_values) != len(col_values):
        return False

    matches = 0

    for row_value, col_value in zip(row_values, col_values):
        if row_value and col_value and row_value == col_value:
            matches += 1

    # Не требуем 100% совпадения, потому что PDF иногда портит отдельные
    # ячейки.
    return matches >= max(2, int(len(col_values) * 0.7))


def convert_value(value):
    """
    Преобразует русские числовые строки в настоящие числа для Excel.

    Примеры:
    "1 234,56"  -> 1234.56
    "12 000"    -> 12000
    "0,25"      -> 0.25
    "12,5%"     -> 12.5
    "100 %"     -> 100

    Текст оставляет без изменений.
    """
    if pd.isna(value):
        return None

    # Если это уже число, оставляем как есть
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()

    if text == "":
        return None

    # Убираем неразрывные пробелы, которые часто бывают в PDF
    text = text.replace("\xa0", " ")
    text = text.replace("\u202f", " ")

    # Минус может быть не обычным дефисом
    text = text.replace("−", "-")

    # Убираем пробелы по краям после нормализации
    text = text.strip()

    # Если значение заканчивается на %, убираем знак процента.
    # В Excel сохраняем именно 12.5, а не 0.125.
    if text.endswith("%"):
        text = text[:-1].strip()

    # Проверяем, похоже ли значение на число:
    # -12 345,67
    # 12 345
    # 123,45
    # 12,5%
    # 100 %
    number_pattern = r"^-?\d{1,3}(?:\s\d{3})*(?:,\d+)?$|^-?\d+(?:,\d+)?$"

    if not re.match(number_pattern, text):
        return value

    # Убираем пробелы-разделители тысяч
    normalized = text.replace(" ", "")

    # Заменяем русскую десятичную запятую на точку
    normalized = normalized.replace(",", ".")

    try:
        number = float(normalized)

        # Если дробной части нет, сохраняем как int
        if number.is_integer():
            return int(number)

        return number

    except ValueError:
        return value


def extract_table_titles_by_page(pdf_path: Path) -> dict[int, list[str]]:
    """
    Возвращает словарь:
    {
        1: ["Таблица 1.1 - ..."],
        2: ["Таблица 1.2 - ..."],
    }

    Ключ — номер страницы PDF, начиная с 1.
    """
    titles_by_page = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""

            titles = TABLE_TITLE_PATTERN.findall(text)

            if titles:
                titles_by_page[page_number] = titles

    return titles_by_page


def write_table_to_sheet(
    ws,
    df: pd.DataFrame,
    start_row: int,
    write_header: bool = True
) -> int:
    """
    Записывает DataFrame на лист Excel.

    Если write_header=True — пишет заголовок колонок.
    Если write_header=False — не пишет заголовок колонок и удаляет
    повторяющуюся строку шапки, если она попала в данные.
    """
    df = df.copy()

    # Удаляем полностью пустые строки
    df = df.dropna(how="all")

    if df.empty:
        return start_row

    # Если это продолжение таблицы, а первая строка повторяет заголовки,
    # удаляем ее.
    if not write_header and len(df) > 0:
        first_row = df.iloc[0].tolist()

        if is_repeated_header_row(first_row, df.columns):
            df = df.iloc[1:].reset_index(drop=True)

    if df.empty:
        return start_row

    current_row = start_row

    # Заголовки таблицы пишем только один раз
    if write_header:
        for col_idx, column_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        current_row += 1

    # Данные таблицы
    for row in df.itertuples(index=False):
        for col_idx, value in enumerate(row, start=1):
            converted_value = convert_value(value)

            cell = ws.cell(
                row=current_row,
                column=col_idx,
                value=converted_value
            )

            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if isinstance(converted_value, int):
                cell.number_format = "#,##0"
            elif isinstance(converted_value, float):
                cell.number_format = "#,##0.00"

        current_row += 1

    # Без пустой строки между продолжениями таблицы
    return current_row


def auto_adjust_columns(ws):
    """
    Простая автоширина колонок.
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value is not None:
                value_length = len(str(cell.value))
                max_length = max(max_length, value_length)

        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)


def convert_pdf_to_excel(pdf_path: Path):
    print(f"Processing: {pdf_path.name}")

    titles_by_page = extract_table_titles_by_page(pdf_path)

    if not titles_by_page:
        print("Не найдено заголовков вида 'Таблица 1.1 - ...'")
        print("Все таблицы будут записаны на один лист.")

    output_file = OUTPUT_DIR / f"{pdf_path.stem}.xlsx"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        workbook = writer.book

        current_sheet_name = "Tables"
        current_ws = workbook.create_sheet(current_sheet_name)
        current_row = 1

        # Удаляем дефолтный пустой лист, если он есть
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        used_sheet_names = set(workbook.sheetnames)

        # Обрабатываем PDF постранично
        # Так проще связать найденную подпись таблицы с таблицами на этой
        # странице
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)

        for page_number in range(1, total_pages + 1):
            page_titles = titles_by_page.get(page_number, [])

            # Если на странице есть подпись "Таблица 1.1 - ...",
            # создаем новый лист.
            # Если подписей несколько, берем первую.
            if page_titles:
                title = page_titles[0]
                base_sheet_name = safe_sheet_name(title)

                sheet_name = base_sheet_name
                counter = 2

                while sheet_name in used_sheet_names:
                    suffix = f"_{counter}"
                    sheet_name = base_sheet_name[: 31 - len(suffix)] + suffix
                    counter += 1

                current_sheet_name = sheet_name
                current_ws = workbook.create_sheet(current_sheet_name)
                used_sheet_names.add(current_sheet_name)
                current_row = 1

                # Пишем название таблицы в первую строку листа
                title_cell = current_ws.cell(row=current_row, column=1,
                                             value=title)
                title_cell.font = Font(bold=True)
                title_cell.alignment = Alignment(wrap_text=True,
                                                 vertical="top")
                current_row += 2

            # Вытаскиваем таблицы только с текущей страницы
            try:
                tables = tabula.read_pdf(
                    str(pdf_path),
                    pages=page_number,
                    multiple_tables=True,
                    lattice=True
                )
            except Exception as error:
                print(f"Ошибка на странице {page_number}: {error}")
                continue

            # Если lattice не нашел таблицы, пробуем stream
            if not tables:
                tables = tabula.read_pdf(
                    str(pdf_path),
                    pages=page_number,
                    multiple_tables=True,
                    stream=True
                )

            if not tables:
                continue

            for table_index, table in enumerate(tables, start=1):
                if table is None or table.empty:
                    continue

                write_header = current_row <= 3

                current_row = write_table_to_sheet(
                    ws=current_ws,
                    df=table,
                    start_row=current_row,
                    write_header=write_header
                )

        # Если лист Tables остался пустым, удаляем его
        if "Tables" in workbook.sheetnames:
            ws = workbook["Tables"]
            if (
                ws.max_row == 1 and ws.max_column == 1
                and ws["A1"].value is None
            ):
                del workbook["Tables"]

        # Настраиваем ширину колонок
        for ws in workbook.worksheets:
            auto_adjust_columns(ws)

    print(f"Saved: {output_file}")


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    pdf_files = list(INPUT_DIR.glob("*.pdf"))

    if not pdf_files:
        print("No PDF files found in input/")
        return

    for pdf_file in pdf_files:
        convert_pdf_to_excel(pdf_file)


if __name__ == "__main__":
    main()
